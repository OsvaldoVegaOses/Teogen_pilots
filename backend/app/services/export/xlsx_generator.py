from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook

from .executive_framework import build_executive_framework
from .privacy import redact_pii_text


class XlsxGenerator:
    @staticmethod
    def _as_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (str, int, float, bool)):
            return redact_pii_text(str(value))
        if isinstance(value, list):
            return " | ".join([XlsxGenerator._as_text(v) for v in value if v is not None])
        if isinstance(value, dict):
            for key in ("text", "description", "definition", "name", "id"):
                if key in value and value[key] is not None:
                    return redact_pii_text(str(value[key]))
            return redact_pii_text(str(value))
        return redact_pii_text(str(value))

    @staticmethod
    def _append_kv_sheet(ws, data: Dict[str, Any]) -> None:
        ws.append(["Campo", "Valor"])
        for key, value in data.items():
            ws.append([key, XlsxGenerator._as_text(value)])

    @staticmethod
    def _as_rows(items: Iterable[Any], limit: int = 200) -> List[str]:
        rows: List[str] = []
        for item in items:
            text = XlsxGenerator._as_text(item).strip()
            if text:
                rows.append(text)
            if len(rows) >= limit:
                break
        return rows

    def generate(self, project_name: str, theory_data: Dict[str, Any], template_key: str = "generic") -> BytesIO:
        wb = Workbook()

        model = theory_data.get("model_json", {}) or {}
        validation = theory_data.get("validation", {}) or {}
        summary = validation.get("network_metrics_summary", {}) or {}
        counts = summary.get("counts", {}) or {}

        ws_summary = wb.active
        ws_summary.title = "Resumen"
        ws_summary.append(["Proyecto", project_name])
        ws_summary.append(["Template", template_key])
        ws_summary.append(["Version", theory_data.get("version", 1)])
        ws_summary.append(["Confianza", theory_data.get("confidence_score", "")])
        ws_summary.append(["Generado por", theory_data.get("generated_by", "TheoGen")])
        ws_summary.append(["Categorias (grafo)", counts.get("category_count", 0)])
        ws_summary.append(["Codigos (grafo)", counts.get("code_count", 0)])
        ws_summary.append(["Fragmentos (grafo)", counts.get("fragment_count", 0)])

        ws_categories = wb.create_sheet("Categorias")
        ws_categories.append(["id", "name", "pagerank", "gds_degree", "code_degree", "fragment_degree"])
        for row in summary.get("category_centrality_top", []) or []:
            ws_categories.append([
                row.get("category_id", ""),
                row.get("category_name", ""),
                row.get("pagerank", ""),
                row.get("gds_degree", ""),
                row.get("code_degree", ""),
                row.get("fragment_degree", ""),
            ])

        ws_co = wb.create_sheet("Coocurrencia")
        ws_co.append(["category_a_id", "category_a_name", "category_b_id", "category_b_name", "shared_fragments"])
        for row in summary.get("category_cooccurrence_top", []) or []:
            ws_co.append([
                row.get("category_a_id", ""),
                row.get("category_a_name", ""),
                row.get("category_b_id", ""),
                row.get("category_b_name", ""),
                row.get("shared_fragments", ""),
            ])

        ws_evidence = wb.create_sheet("Evidencia")
        ws_evidence.append(["category_id", "category_name", "fragment_id", "score", "text", "codes"])
        for bucket in summary.get("semantic_evidence_top", []) or []:
            cid = bucket.get("category_id", "")
            cname = bucket.get("category_name", "")
            for frag in bucket.get("fragments", []) or []:
                ws_evidence.append([
                    cid,
                    cname,
                    frag.get("fragment_id") or frag.get("id", ""),
                    frag.get("score", ""),
                    self._as_text(frag.get("text", "")),
                    self._as_text(frag.get("codes", [])),
                ])

        ws_paradigm = wb.create_sheet("Paradigma")
        self._append_kv_sheet(
            ws_paradigm,
            {
                "selected_central_category": model.get("selected_central_category", ""),
                "conditions": model.get("conditions", ""),
                "actions": model.get("actions", ""),
                "consequences": model.get("consequences", ""),
                "context": model.get("context", ""),
                "intervening_conditions": model.get("intervening_conditions", ""),
            },
        )

        ws_props = wb.create_sheet("Proposiciones")
        ws_props.append(["#", "texto"])
        for idx, text in enumerate(self._as_rows(theory_data.get("propositions", []) or []), start=1):
            ws_props.append([idx, text])

        ws_gaps = wb.create_sheet("Brechas")
        ws_gaps.append(["#", "descripcion"])
        for idx, text in enumerate(self._as_rows(theory_data.get("gaps", []) or []), start=1):
            ws_gaps.append([idx, text])

        ws_metrics = wb.create_sheet("Metricas")
        ws_metrics.append(["metric", "value"])
        ws_metrics.append(["confidence_score", theory_data.get("confidence_score", "")])
        ws_metrics.append(["propositions_count", len(theory_data.get("propositions", []) or [])])
        ws_metrics.append(["gaps_count", len(theory_data.get("gaps", []) or [])])
        ws_metrics.append(["network_category_count", counts.get("category_count", 0)])
        ws_metrics.append(["network_code_count", counts.get("code_count", 0)])
        ws_metrics.append(["network_fragment_count", counts.get("fragment_count", 0)])

        framework = build_executive_framework(theory_data)
        quality = framework.get("quality_metrics", {}) if isinstance(framework.get("quality_metrics"), dict) else {}

        ws_decision = wb.create_sheet("DecisionFramework")
        ws_decision.append(["Campo", "Valor"])
        ws_decision.append(["recommendation", framework.get("recommendation", "PILOT")])
        ws_decision.append(["claims_count", quality.get("claims_count", 0)])
        ws_decision.append(["claims_without_evidence", quality.get("claims_without_evidence", 0)])
        ws_decision.append(["interviews_covered", quality.get("interviews_covered", 0)])
        ws_decision.append(["gaps_count", quality.get("gaps_count", 0)])
        ws_decision.append(["gaps_high", quality.get("gaps_high", 0)])
        ws_decision.append(["judge_warn_only", quality.get("judge_warn_only", False)])
        ws_decision.append([])
        ws_decision.append(["Razon"])
        for reason in self._as_rows(framework.get("reasons", []), limit=20):
            ws_decision.append([reason])

        ws_action = wb.create_sheet("PlanAccion")
        ws_action.append(["#", "accion"])
        for idx, action in enumerate(self._as_rows(framework.get("action_plan", []), limit=20), start=1):
            ws_action.append([idx, action])

        ws_limits = wb.create_sheet("Limites")
        ws_limits.append(["#", "limitacion"])
        for idx, item in enumerate(self._as_rows(framework.get("limitations", []), limit=20), start=1):
            ws_limits.append([idx, item])

        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    cell_value = "" if cell.value is None else str(cell.value)
                    if len(cell_value) > max_len:
                        max_len = len(cell_value)
                ws.column_dimensions[col_letter].width = min(80, max(12, max_len + 2))

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
